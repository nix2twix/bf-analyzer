# === LIBRARIES GENERAL ===
import io
import os
import re
import cv2
import json
import zipfile

import numpy as np
import pandas as pd
import xml.etree.ElementTree as ET

from PIL import Image
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, 
    PatternFill, 
    Alignment
)
from openpyxl.utils import get_column_letter


def rle_encode(mask: np.ndarray, left: int, top: int, right: int, bottom: int):
    pixels = mask.flatten(order="C")
    diff = np.diff(pixels)
    change_idx = np.where(diff != 0)[0] + 1
    splits = np.split(pixels, change_idx)

    result = [(block[0], len(block)) for block in splits]

    rle = []
    idx = 0
    for val, count in result:
        if idx == 0:
            if val == 1:  
                rle.append(0)              
                rle.append(float(count))  
            else:
                rle.append(float(count))
            idx = 1
        else:
            rle.append(count)

    # bbox
    rle.extend([float(left), float(top), float(right), float(bottom)])
    return rle


def rle_decode(rle_str, shape):
    rle_numbers = [int(num) for num in re.findall(r'\d+', rle_str)]
    img = np.zeros(shape[0] * shape[1], dtype=np.uint8)

    index = 0
    i = 0
    n = len(rle_numbers)

    while i < n:
        start = rle_numbers[i]
        i += 1
        length = 0
        if i < n:
            length = rle_numbers[i]
            i += 1
        index += start
        if length > 0:
            if index + length > img.size:
                raise ValueError(f"RLE segment exceeds mask size: index {index} length {length} size {img.size}")
            img[index:index+length] = 1
            index += length
    return img.reshape(shape)

def loadMasksFromZip(uploaded_zip, imgWidth, imgHeight, model_config):
    if uploaded_zip is None:
        return None

    try:
        with zipfile.ZipFile(uploaded_zip) as z:
            xml_filename = next((name for name in z.namelist() if name.lower().endswith('.xml')), None)
            if xml_filename is None:
                print("No XML file found in the uploaded ZIP archive.")
                return None
            
            with z.open(xml_filename) as xml_file:
                xml_bytes = xml_file.read()
                predicted_objects = updateMaskCVAT(xml_bytes, imgWidth, imgHeight, model_config)
                return predicted_objects

    except zipfile.BadZipFile:
        print("Uploaded file is not a valid ZIP archive.")
        return None


def updateMaskCVAT(xml_path, width, height, model_config):
    """
    Обновление масок из CVAT XML с использованием конфига модели
    """
    # Строим обратный маппинг: cvat_label -> internal_class
    label_map = {}
    if model_config and model_config.cvat_labels:
        for internal_class, cvat_label in model_config.cvat_labels.items():
            label_map[cvat_label] = internal_class
    
    print(f"[DEBUG] updateMaskCVAT label_map: {label_map}")
    
    xml_io = io.BytesIO(xml_path)
    tree = ET.parse(xml_io)
    root = tree.getroot()

    # Инициализируем маски для всех классов кроме фона
    masks = {}
    next_obj_id = {}
    for class_name in model_config.class_labels.keys():
        if "background" not in class_name and class_name != "bg":
            masks[class_name] = np.zeros((height, width), dtype=np.uint16)
            next_obj_id[class_name] = 1

    for image in root.findall('image'):
        for obj in list(image):
            label = obj.attrib.get('label')
            if label not in label_map:
                print(f"[DEBUG] Skipping unknown label: {label}")
                continue

            class_key = label_map[label]
            print(f"[DEBUG] Processing {label} -> {class_key}")

            if obj.tag == 'mask':
                # ==== RLE ====
                rle = obj.attrib['rle']
                left = int(obj.attrib.get('left', 0))
                top = int(obj.attrib.get('top', 0))
                w = int(obj.attrib.get('width', 0))
                h = int(obj.attrib.get('height', 0))

                decoded = rle_decode(rle, (h, w)).astype(np.uint8)
                if np.any(decoded):
                    obj_id = next_obj_id[class_key]
                    masks[class_key][top:top+h, left:left+w][decoded > 0] = obj_id
                    next_obj_id[class_key] += 1
                    print(f"[DEBUG]   Added mask object {obj_id}")

            elif obj.tag == 'polygon':
                # ==== Полигоны ====
                points_str = obj.attrib['points']
                pts = []
                for p in points_str.split(';'):
                    x, y = map(float, p.split(','))
                    pts.append([int(round(x)), int(round(y))])
                pts = np.array(pts, dtype=np.int32)

                obj_id = next_obj_id[class_key]
                cv2.fillPoly(masks[class_key], [pts], color=obj_id)
                next_obj_id[class_key] += 1
                print(f"[DEBUG]   Added polygon object {obj_id}")

    # Выводим статистику
    for class_name, mask in masks.items():
        unique_ids = np.unique(mask)
        obj_count = len(unique_ids[unique_ids != 0])
        print(f"[DEBUG] Class {class_name}: {obj_count} objects loaded")

    return masks

def makeCVATbackupRLE(
    image: Image.Image,
    original_filename: str,
    filtered_objects: dict,
    width: int,
    height: int,
    model_config 
):
    """
    Создаёт ZIP-архив для CVAT на основе конфига модели
    """
    # Используем cvat_labels из конфига
    class_to_cvat = model_config.cvat_labels or {}
    
    # Получаем список всех меток для task.json
    all_cvat_labels = []
    for cvat_name, color in model_config.cvat_label_colors.items():
        all_cvat_labels.append({
            "name": cvat_name,
            "color": color,
            "attributes": [],
            "type": "any",
            "sublabels": []
        })

    base_name, ext = os.path.splitext(original_filename)
    archive_name = f"backup-{datetime.now().strftime('%d-%m-%Y-%H-%M')}.zip"

    # manifest.jsonl
    manifest = [
        {"version": "1.1"},
        {"type": "images"},
        {"name": base_name, "extension": ext, "width": width, "height": height, "meta": {"related_images": []}},
    ]
    manifest_content = "\n".join(json.dumps(line, ensure_ascii=False) for line in manifest)

    # task.json
    task = {
        "name": archive_name,
        "bug_tracker": "",
        "status": "annotation",
        "subset": "Train",
        "labels": all_cvat_labels,  
        "version": "1.0",
        "data": {
            "chunk_size": 15,
            "image_quality": 100,
            "start_frame": 0,
            "stop_frame": 0,
            "storage_method": "cache",
            "storage": "local",
            "sorting_method": "lexicographical",
            "chunk_type": "imageset",
            "deleted_frames": [],
        },
        "jobs": [{"status": "annotation", "type": "annotation", "start_frame": 0, "stop_frame": 0, "frames": []}],
    }
    task_content = json.dumps(task, ensure_ascii=False, indent=2)

    # annotations.json
    frame_obj = {"version": 0, "tags": [], "shapes": [], "tracks": []}

    for class_name, labeled_mask in filtered_objects.items():
        if class_name not in class_to_cvat:
            continue 

        cvat_label = class_to_cvat[class_name]
        unique_ids = np.unique(labeled_mask)
        unique_ids = unique_ids[unique_ids != 0]

        for obj_id in unique_ids:
            obj_mask = (labeled_mask == obj_id).astype(np.uint8)
            ys, xs = np.where(obj_mask > 0)
            
            if len(ys) == 0 or len(xs) == 0:
                continue
                
            top, left = ys.min(), xs.min()
            bottom, right = ys.max(), xs.max()
            w, h = right - left + 1, bottom - top + 1

            crop = obj_mask[top:bottom + 1, left:right + 1]
            points = rle_encode(crop, left, top, right, bottom)
            
            frame_obj["shapes"].append({
                "type": "mask",
                "occluded": False,
                "outside": False,
                "z_order": 0,
                "rotation": 0.0,
                "points": points,
                "frame": 0,
                "group": 0,
                "source": "manual",
                "attributes": [],
                "label": cvat_label,
            })
    
    annotations = [frame_obj]
    annotations_content = json.dumps(annotations, ensure_ascii=False, separators=(',', ':'))
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        img_bytes = io.BytesIO()
        image.save(img_bytes, format=image.format or ext.replace(".", "").upper())
        img_bytes.seek(0)
        zf.writestr(f"data/{original_filename}", img_bytes.read())
        zf.writestr("data/manifest.jsonl", manifest_content)
        zf.writestr("task.json", task_content)
        zf.writestr("annotations.json", annotations_content)

    zip_buffer.seek(0)
    return zip_buffer


def makeXMLforCVAT(image_name, image_width, image_height, labels, masks, crop_bottom=120):
    """
    labels: [{"name": str, "color": str}]
    masks: [{"label": str, "mask": np.ndarray, "z_order": int}]
    crop_bottom: сколько пикселей снизу было обрезано (для восстановления)
    """
    root = ET.Element("annotations")
    ET.SubElement(root, "version").text = "1.1"

    meta = ET.SubElement(root, "meta")
    job = ET.SubElement(meta, "job")
    ET.SubElement(job, "id").text = "1"
    ET.SubElement(job, "size").text = "1"
    ET.SubElement(job, "mode").text = "annotation"
    ET.SubElement(job, "overlap").text = "0"
    ET.SubElement(job, "created").text = datetime.utcnow().isoformat()
    ET.SubElement(job, "updated").text = datetime.utcnow().isoformat()
    labels_el = ET.SubElement(job, "labels")
    for lbl in labels:
        lbl_el = ET.SubElement(labels_el, "label")
        ET.SubElement(lbl_el, "name").text = lbl["name"]
        ET.SubElement(lbl_el, "color").text = lbl["color"]
        ET.SubElement(lbl_el, "type").text = "any"
        ET.SubElement(lbl_el, "attributes").text = " "

    img_el = ET.SubElement(root, "image", {
        "id": "0",
        "name": image_name,
        "width": str(image_width),
        "height": str(image_height)
    })

    for m in masks:
        # 1. Восстанавливаем высоту, если маска обрезана
        mask_array = m["mask"].astype(np.uint8)
        if mask_array.shape[0] != image_height:
            pad_height = image_height - mask_array.shape[0]
            if pad_height < 0:
                raise ValueError("Маска больше, чем изображение!")
            mask_array = np.pad(mask_array, ((0, pad_height), (0, 0)), mode='constant', constant_values=0)

        # 2. Разделяем на объекты
        num_labels, labels_im = cv2.connectedComponents(mask_array)
        for obj_id in range(1, num_labels):  # 0 — фон
            obj_mask = (labels_im == obj_id).astype(np.uint8)

            # 3. Bounding box
            ys, xs = np.where(obj_mask > 0)
            if len(xs) == 0:
                continue
            left, top = xs.min(), ys.min()
            width = xs.max() - xs.min() + 1
            height = ys.max() - ys.min() + 1

            cropped_mask = obj_mask[top:top+height, left:left+width]
            rle = rle_encode(cropped_mask)

            ET.SubElement(img_el, "mask", {
                "label": m["label"],
                "source": "manual",
                "occluded": "0",
                "rle": rle,
                "left": str(left),
                "top": str(top),
                "width": str(width),
                "height": str(height),
                "z_order": str(m.get("z_order", 0))
            }).text = " "

    return ET.tostring(root, encoding="utf-8", xml_declaration=True).decode("utf-8")

def saveResultsAsZip(filteredObjects, classColors, drawImgPIL, 
                     filteredObjectsInfo=None, scale=None, 
                     imgWidth=None, imgHeight=None, infoLineHeight=None,
                     filtration_params=None, resultInfo=None, imgName=None,
                     model_config=None):
    """
    Сохраняет результаты в ZIP-архив с изображениями и Excel отчетом
    
    Args:
        filteredObjects: словарь с масками объектов по классам
        classColors: словарь цветов для классов
        drawImgPIL: PIL изображение для визуализации
        filteredObjectsInfo: DataFrame или список с информацией об объектах
        scale: масштаб изображения (мкм/пиксель)
        imgWidth: ширина исходного изображения
        imgHeight: высота исходного изображения
        infoLineHeight: высота информационной строки
        filtration_params: параметры фильтрации
        resultInfo: словарь с результатами статистики (count, total_area_mkm и т.д.)
        imgName: имя изображения
    """
    drawImg = np.array(drawImgPIL)
    if drawImg.ndim == 2: 
        drawImg = cv2.cvtColor(drawImg, cv2.COLOR_GRAY2RGBA)
    else:
        drawImg = cv2.cvtColor(drawImg, cv2.COLOR_RGB2RGBA)
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, mode="w") as zf:
        buf = io.BytesIO()
        drawImgPIL.save(buf, format="PNG")
        zf.writestr("segmented_image.png", buf.getvalue())
        
        h, w = drawImg.shape[:2]
        all_objects_mask = np.zeros((h, w, 4), dtype=np.uint8)
        objects_stats = []
        
        for className, class_mask in filteredObjects.items():
            if className not in classColors or className == "bg":
                continue

            class_color = classColors[className]
            if len(class_color) == 3:
                class_color = (*class_color, 255) 
                
            class_rgba = np.zeros((h, w, 4), dtype=np.uint8)

            obj_ids = [id for id in np.unique(class_mask) if id != 0]
            for obj_id in obj_ids:
                obj_mask = (class_mask == obj_id).astype(np.uint8)
                area = np.sum(obj_mask)
                contours, _ = cv2.findContours(
                    obj_mask,
                    cv2.RETR_EXTERNAL,
                    cv2.CHAIN_APPROX_SIMPLE
                )
                
                perimeter = 0
                if contours:
                    perimeter = cv2.arcLength(contours[0], True)
                
                M = cv2.moments(obj_mask)
                if M["m00"] > 0:
                    cx = int(M["m10"] / M["m00"])
                    cy = int(M["m01"] / M["m00"])
                else:
                    cx, cy = 0, 0
                
                objects_stats.append({
                    'Класс': className,
                    'ID объекта': obj_id,
                    'Площадь (пикс)': area,
                    f'Площадь (мкм²)': area * (scale ** 2) if scale else 0,
                    'Периметр (пикс)': perimeter,
                    f'Периметр (мкм)': perimeter * scale if scale else 0,
                    'Центр X': cx,
                    'Центр Y': cy
                })

                class_rgba[obj_mask > 0] = class_color
                all_objects_mask[obj_mask > 0] = class_color

                if contours:
                    cv2.drawContours(
                        class_rgba,
                        contours,
                        -1,
                        (255, 255, 255, 255),
                        2
                    )
                    
                    cv2.drawContours(
                        all_objects_mask,
                        contours,
                        -1,
                        (255, 255, 255, 255),
                        2
                    )
                
                if M["m00"] > 0:
                    cv2.putText(
                        class_rgba,
                        str(obj_id),
                        (cx, cy),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        1.2,
                        (255, 255, 255, 255),
                        2
                    )
                    cv2.putText(
                        all_objects_mask,
                        str(obj_id),
                        (cx, cy),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        1.2,
                        (255, 255, 255, 255),
                        2
                    )
            
            if np.any(class_rgba):
                buf = io.BytesIO()
                Image.fromarray(class_rgba).save(buf, format="PNG")
                zf.writestr(f"{className}_mask.png", buf.getvalue())
        
        if np.any(all_objects_mask):
            buf = io.BytesIO()
            Image.fromarray(all_objects_mask).save(buf, format="PNG")
            zf.writestr("all_objects_mask.png", buf.getvalue())
        
        # ===== EXCEL =====
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            file_info = pd.DataFrame({
                'Параметр': [
                    'Имя файла',
                    'Дата обработки',
                    'Ширина изображения (пикс)',
                    'Высота изображения (пикс)',
                    'Высота информационной строки (пикс)',
                    'Масштаб (мкм/пикс)',
                    'Площадь изображения (мкм²)'
                ],
                'Значение': [
                    imgName or 'Не указано',
                    pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                    imgWidth or '—',
                    imgHeight or '—',
                    infoLineHeight or '0',
                    scale or '—',
                    f"{imgWidth * (imgHeight - infoLineHeight) * (scale ** 2):.2f}" 
                    if all([imgWidth, imgHeight, scale]) else '—'
                ]
            })
            file_info.to_excel(writer, sheet_name='Информация о файле', index=False)
            
            if filtration_params and model_config:
                params_list = []
                for param_name, (min_val, max_val) in model_config.filtration_params.items():
                        if param_name in filtration_params:
                            current = filtration_params[param_name]
                            # Если current это словарь с min/max
                            if isinstance(current, dict):
                                value_display = f"{current['min']:.1f} - {current['max']:.1f}"
                            else:
                                value_display = f"{current:.1f}"
            
                            # Красивое имя для отображения
                            display_name = param_name.replace("_", " ").title()
            
                            params_list.append({
                                'Параметр': display_name,
                                'Значение': value_display
                            })
    
                if params_list:
                    params_df = pd.DataFrame(params_list)
                    params_df.to_excel(writer, sheet_name='Параметры фильтрации', index=False)
            
            if resultInfo:
                img_area_um2 = 0
                if all([imgWidth, imgHeight, infoLineHeight, scale]):
                    img_area_um2 = imgWidth * (imgHeight - infoLineHeight) * (scale ** 2)
                
                summary_data = []
                class_titles = {}
                if model_config and hasattr(model_config, 'class_titles'):
                    for class_name, (title, color) in model_config.class_titles.items():
                        class_titles[class_name] = title
                
                for class_name, class_title in class_titles.items():
                    if class_name in resultInfo:
                        stats = resultInfo[class_name]
                        area_percent = (stats['total_area_mkm'] / img_area_um2 * 100) if img_area_um2 > 0 else 0
                        
                        summary_data.append({
                            'Класс': class_title,
                            'Количество объектов': stats['count'],
                            'Общая площадь (мкм²)': f"{stats['total_area_mkm']:.2f}",
                            'Процент от площади': f"{area_percent:.2f}%"
                        })
                
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Общая статистика', index=False)
            
            if objects_stats:
                objects_df = pd.DataFrame(objects_stats)
    
                if all([imgWidth, imgHeight, infoLineHeight, scale]) and resultInfo:
                    img_area_um2 = imgWidth * (imgHeight - infoLineHeight) * (scale ** 2)
        
                    class_total_area = {}
                    for class_name, stats in resultInfo.items():
                        if isinstance(stats, dict) and 'total_area_mkm' in stats:
                            class_total_area[class_name] = stats['total_area_mkm']
        
                    objects_df['% от площади класса'] = objects_df.apply(
                        lambda row: f"{(row['Площадь (мкм²)'] / class_total_area.get(row['Класс'], 1) * 100):.2f}%" 
                        if row['Класс'] in class_total_area else '—',
                        axis=1
                    )
                    objects_df['% от общей площади'] = objects_df.apply(
                        lambda row: f"{(row['Площадь (мкм²)'] / img_area_um2 * 100):.2f}%",
                        axis=1
                    )
    
                objects_df.to_excel(writer, sheet_name='Объекты', index=False)
                
                objects_df.to_excel(writer, sheet_name='Объекты', index=False)
            
            # # 5. ЛИСТ: Информация из filteredObjectsInfo (если есть)
            # if filteredObjectsInfo is not None:
            #     try:
            #         if isinstance(filteredObjectsInfo, str):
            #             from io import StringIO
            #             info_df = pd.read_csv(StringIO(filteredObjectsInfo))
            #         elif isinstance(filteredObjectsInfo, pd.DataFrame):
            #             info_df = filteredObjectsInfo
            #         else:
            #             info_df = pd.DataFrame(filteredObjectsInfo)
                    
            #         if not info_df.empty:
            #             # Если данных много, оставляем как есть
            #             info_df.to_excel(writer, sheet_name='Доп. информация', index=False)
            #     except Exception as e:
            #         # Если не удалось распарсить, сохраняем как текст в транспонированном виде
            #         text_df = pd.DataFrame({
            #             'Параметр': ['Дополнительная информация'],
            #             'Значение': [str(filteredObjectsInfo)[:32767]]  # Ограничение Excel
            #         })
            #         text_df.to_excel(writer, sheet_name='Доп. информация', index=False)
        

        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
        
        formatted_excel_buffer = io.BytesIO()
        wb.save(formatted_excel_buffer)
        formatted_excel_buffer.seek(0)
        zf.writestr(f"results-{datetime.now().strftime('%Y-%m-%d-%H-%M')}.xlsx", formatted_excel_buffer.getvalue())
    
    zip_buffer.seek(0)
    return zip_buffer